#!/usr/bin/env python3
"""V2 ultra-short metric evaluator using index0 groundtruth as the target.

This version keeps the time-alignment and metric implementation from
``evaluate_ultrashort_metric.py`` but changes the target-value policy: after
all forecast horizons are aligned to their target timestamps, the groundtruth
from index0 is used as the sole groundtruth for every horizon at that target.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

import evaluate_ultrashort_metric as v1


GROUNDTRUTH_TOLERANCE = 1e-6


def apply_index0_groundtruth(
    records: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Replace every horizon's target with index0 groundtruth at the same time."""
    required = {"target_timestamp", "forecast_index", "groundtruth"}
    missing = required.difference(records.columns)
    if missing:
        raise KeyError(f"Records are missing columns: {sorted(missing)}")

    index0 = records.loc[
        records["forecast_index"] == 0, ["target_timestamp", "groundtruth"]
    ].copy()
    if index0.empty:
        raise ValueError(
            "V2 requires an index0 parquet file because index0 groundtruth is the target"
        )
    duplicate = index0.duplicated("target_timestamp", keep=False)
    if duplicate.any():
        sample = index0.loc[duplicate, "target_timestamp"].iloc[0]
        raise ValueError(f"index0 contains duplicate target timestamp {sample}")

    index0_target = index0.set_index("target_timestamp")["groundtruth"]
    result = records.copy()
    result["index0_groundtruth"] = result["target_timestamp"].map(index0_target)

    comparable = result["groundtruth"].notna() & result["index0_groundtruth"].notna()
    conflict = comparable & (
        (result["groundtruth"] - result["index0_groundtruth"]).abs()
        > GROUNDTRUTH_TOLERANCE
    )
    conflict_target_count = int(result.loc[conflict, "target_timestamp"].nunique())
    missing_index0_target_count = int(
        result.loc[result["index0_groundtruth"].isna(), "target_timestamp"].nunique()
    )

    # Do not fall back to another horizon.  If index0 has no target for a time,
    # the target remains missing and the base evaluator applies its missing rule.
    result["groundtruth"] = result["index0_groundtruth"]
    result = result.drop(columns="index0_groundtruth")
    return result, {
        "conflict_target_count": conflict_target_count,
        "missing_index0_target_count": missing_index0_target_count,
    }


def write_excel_report_v2(
    daily: pd.DataFrame,
    monthly: pd.DataFrame,
    output_path: Path,
    *,
    capacity: float,
    input_files: list[Path],
    missing_normalized_error: float,
    complete_days_only: bool,
    target_stats: dict[str, int],
) -> None:
    """Write the normal report and document the V2 target-selection policy."""
    v1.write_excel_report(
        daily,
        monthly,
        output_path,
        capacity=capacity,
        input_files=input_files,
        missing_normalized_error=missing_normalized_error,
        complete_days_only=complete_days_only,
    )

    from openpyxl import load_workbook
    from openpyxl.styles import Font

    workbook = load_workbook(output_path)
    sheet = workbook["计算说明"]
    extra_rows = [
        ("脚本版本", "v2"),
        (
            "真值策略",
            "同一目标时刻统一采用 index0（起报时间+15分钟）的 groundtruth",
        ),
        ("发现真值冲突的目标时刻数", target_stats["conflict_target_count"]),
        ("index0 缺少真值的目标时刻数", target_stats["missing_index0_target_count"]),
    ]
    for label, value in extra_rows:
        sheet.append([label, value])
        sheet.cell(sheet.max_row, 1).font = Font(bold=True)
    sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width, 42)
    workbook.save(output_path)


def build_argument_parser():
    parser = v1.build_argument_parser()
    parser.description = (
        "Calculate the revised ultra-short metric using index0 groundtruth as target"
    )
    parser.set_defaults(output=Path("forecast_metrics_v2.xlsx"))
    return parser


def main() -> None:
    args = build_argument_parser().parse_args()
    files = v1.discover_input_files(args.input_glob)
    records = v1.load_forecast_records(
        files,
        timestamp_column=args.timestamp_column,
        prediction_column=args.prediction_column,
        groundtruth_column=args.groundtruth_column,
    )
    records, target_stats = apply_index0_groundtruth(records)
    daily = v1.calculate_daily_metrics(
        records,
        capacity=args.capacity,
        start_date=args.start_date,
        end_date=args.end_date,
        missing_normalized_error=args.missing_normalized_error,
    )
    monthly = v1.calculate_monthly_metrics(
        daily, complete_days_only=args.monthly_complete_days_only
    )
    write_excel_report_v2(
        daily,
        monthly,
        args.output,
        capacity=args.capacity,
        input_files=files,
        missing_normalized_error=args.missing_normalized_error,
        complete_days_only=args.monthly_complete_days_only,
        target_stats=target_stats,
    )
    print(
        f"Wrote {len(daily)} daily rows and {len(monthly)} monthly rows to "
        f"{args.output}; index0 replaced conflicting groundtruth at "
        f"{target_stats['conflict_target_count']} target timestamps"
    )


if __name__ == "__main__":
    main()
